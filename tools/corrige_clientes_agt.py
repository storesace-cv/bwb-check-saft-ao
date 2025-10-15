"""Ferramentas para corrigir listagens de clientes usando dados da AGT."""
from __future__ import annotations

import json
import math
import os
import re
import time
import unicodedata
from typing import Any, Iterable, Mapping

import pandas as pd
import requests
from openpyxl.styles import Font, PatternFill
from tenacity import RetryError, retry, retry_if_exception_type, stop_after_attempt, wait_exponential

API_BASE_DEFAULT = "https://invoice.minfin.gov.ao/commonServer/common/taxpayer/get/"
NIF_PT_API_BASE_DEFAULT = "http://www.nif.pt/"
NIF_PT_API_KEY_ENV = "NIF_PT_API_KEY"
NIF_PT_API_KEY_DEFAULT = "ccba688830ceb8a4d0f574fb4c7f6df6"
API_TIMEOUT = 10.0
RATE_LIMIT = 5.0
USE_CACHE = True

_LAST_REQUEST_AT = 0.0
LAST_SUMMARY: dict[str, Any] | None = None


class TransientAPIError(Exception):
    """Erro transitório para disparar novos retries."""


def set_fetch_settings(*, rate_limit: float | None = None, timeout: float | None = None, use_cache: bool | None = None) -> None:
    """Permite ajustar definições globais de chamadas à API."""
    global RATE_LIMIT, API_TIMEOUT, USE_CACHE
    if rate_limit is not None:
        RATE_LIMIT = float(rate_limit)
    if timeout is not None:
        API_TIMEOUT = float(timeout)
    if use_cache is not None:
        USE_CACHE = bool(use_cache)


def _throttle_if_needed() -> None:
    global _LAST_REQUEST_AT
    if RATE_LIMIT and RATE_LIMIT > 0:
        min_interval = 1.0 / RATE_LIMIT
        now = time.monotonic()
        elapsed = now - _LAST_REQUEST_AT
        if elapsed < min_interval:
            time.sleep(min_interval - elapsed)
        _LAST_REQUEST_AT = time.monotonic()


def normalizar_nif(raw: Any) -> str:
    """Normaliza o NIF removendo caracteres não alfanuméricos."""
    if raw is None:
        return ""
    if isinstance(raw, float):
        if math.isnan(raw):
            return ""
        if raw.is_integer():
            raw = str(int(raw))
        else:
            raw = str(raw)
    else:
        raw = str(raw)
    return "".join(ch for ch in raw.strip().upper() if ch.isalnum())


def classificar_nif_ao(raw: Any) -> str:
    """Classifica heurísticamente um NIF de Angola."""
    nif = normalizar_nif(raw)
    if not nif:
        return "manifestamente_errado"
    if not re.fullmatch(r"[0-9A-Z]+", nif):
        return "manifestamente_errado"
    if nif[0].isalpha():
        return "manifestamente_errado"
    if nif == "999999999":
        return "possivelmente_correto"
    if re.fullmatch(r"\d{10}", nif):
        return "possivelmente_correto"
    if re.fullmatch(r"\d{9}[A-Z]{2}\d{3}", nif):
        return "possivelmente_correto"
    if re.fullmatch(r"[A-Z0-9]{9,14}", nif):
        return "possivelmente_correto"
    if len(nif) < 6 or len(nif) > 15:
        return "manifestamente_errado"
    return "possivelmente_errado"


def validar_nif_portugal(nif: str) -> tuple[bool, str | None]:
    """Valida um NIF português e devolve o prefixo identificado."""

    if len(nif) != 9 or not nif.isdigit():
        return False, None

    prefixo: str | None
    if nif.startswith("45"):
        prefixo = "45"
    else:
        prefixo = nif[0]
        if prefixo not in {"1", "2", "3", "5", "6", "8", "9"}:
            return False, None

    soma = sum(int(digito) * (9 - idx) for idx, digito in enumerate(nif[:8]))
    resto = soma % 11
    digito_controle = 0 if resto < 2 else 11 - resto
    if digito_controle != int(nif[-1]):
        return False, None

    return True, prefixo


def fetch_taxpayer_pt(
    nif: str,
    session: requests.Session,
    cache: dict[str, dict[str, Any] | None] | None,
) -> dict[str, Any] | None:
    """Obtém dados de contribuinte via NIF.PT."""

    if not nif:
        return None

    cache_enabled = USE_CACHE and cache is not None
    if cache_enabled and nif in cache:
        return cache[nif]

    key = os.getenv(NIF_PT_API_KEY_ENV, NIF_PT_API_KEY_DEFAULT)
    base = os.getenv("NIF_PT_API_BASE", NIF_PT_API_BASE_DEFAULT)
    params = {"json": "1", "q": nif, "key": key}

    try:
        _throttle_if_needed()
        response = session.get(base, params=params, timeout=API_TIMEOUT)
    except requests.RequestException:
        result = None
    else:
        if response.status_code >= 400:
            result = None
        else:
            try:
                payload = response.json()
            except json.JSONDecodeError:
                result = None
            else:
                result = _interpretar_payload_pt(payload)

    if cache_enabled:
        cache[nif] = result
    return result


def _interpretar_payload_pt(payload: Any) -> dict[str, Any] | None:
    """Tenta extrair dados relevantes da resposta do NIF.PT."""

    if not isinstance(payload, Mapping):
        return None

    if payload.get("result") in {"error", False}:
        return None

    candidato: Any = None

    registros = payload.get("records")
    if isinstance(registros, list) and registros:
        candidato = registros[0]
    elif isinstance(registros, Mapping) and registros:
        # Respostas do nif.pt podem vir num dict {"<nif>": {...}}
        primeiro = next(iter(registros.values()))
        if isinstance(primeiro, Mapping):
            candidato = primeiro
        else:
            candidato = None
    elif isinstance(payload.get("data"), Mapping):
        candidato = payload["data"]
    elif isinstance(payload.get("record"), Mapping):
        candidato = payload["record"]
    else:
        candidato = payload

    if not isinstance(candidato, Mapping):
        return None

    nome = (
        candidato.get("name")
        or candidato.get("companyName")
        or candidato.get("nome")
        or candidato.get("title")
    )
    morada = (
        candidato.get("address")
        or candidato.get("morada")
        or candidato.get("addressDetail")
        or candidato.get("address_detail")
        or candidato.get("addressdetail")
    )

    if not nome and not morada:
        return None

    localidade = (
        candidato.get("city")
        or candidato.get("localidade")
        or candidato.get("municipio")
        or candidato.get("county")
    )

    return {"name": nome, "address": morada, "city": localidade}


def _mensagem_nif_invalido(classificacao: str) -> str:
    if classificacao == "manifestamente_errado":
        return "NIF INVALIDO | Manifestamente errado"
    return "NIF INVALIDO | Possivelmente errado"


PORTUGAL_SINGULAR_PREFIXES = {"1", "2", "3"}
PORTUGAL_CORPORATE_PREFIXES = {"45", "5", "6", "8"}
PORTUGAL_PREFIX_MESSAGES = {
    "singular": "NIF INVALIDO | Possivelmente Português (Pessoas singulares)",
    "45": "NIF INVALIDO | Possivelmente Português (Pessoas coletivas não residentes)",
    "5": "NIF INVALIDO | Possivelmente Português (Pessoas coletivas (empresas)",
    "6": "NIF INVALIDO | Possivelmente Português (Administrações públicas)",
    "8": "NIF INVALIDO | Possivelmente Português (Empresários em nome individual)",
}


def avaliar_nif_portugues(
    nif: str,
    session: requests.Session,
    cache: dict[str, dict[str, Any] | None] | None,
) -> dict[str, Any] | None:
    """Analisa se um NIF inválido para Angola é potencialmente português."""

    valido, prefixo = validar_nif_portugal(nif)
    if not valido or prefixo is None:
        return None

    resultado: dict[str, Any] = {
        "pais": "Portugal",
        "mensagem": None,
        "nome": None,
        "morada": None,
        "localidade": None,
        "prefixo": prefixo,
    }

    if prefixo in PORTUGAL_SINGULAR_PREFIXES:
        resultado["mensagem"] = PORTUGAL_PREFIX_MESSAGES["singular"]
        resultado["localidade"] = resultado["mensagem"]
        return resultado

    if prefixo in PORTUGAL_CORPORATE_PREFIXES:
        dados_pt = fetch_taxpayer_pt(nif, session, cache)
        if dados_pt:
            if dados_pt.get("name"):
                resultado["nome"] = dados_pt["name"]
            if dados_pt.get("address"):
                resultado["morada"] = dados_pt["address"]
            if dados_pt.get("city"):
                resultado["localidade"] = dados_pt["city"]
        if not any((resultado["nome"], resultado["morada"], resultado["localidade"])):
            resultado["mensagem"] = PORTUGAL_PREFIX_MESSAGES[prefixo]
        return resultado

    return None


def _build_url(nif: str) -> str:
    base = os.getenv("AGT_TAXPAYER_BASE", API_BASE_DEFAULT)
    if not base.endswith("/"):
        base = f"{base}/"
    return f"{base}{nif}"


def fetch_taxpayer(nif: str, session: requests.Session, cache: dict[str, dict[str, Any] | None] | None) -> dict[str, Any] | None:
    """Obtém dados do contribuinte via API pública da AGT."""
    if not nif:
        return None

    cache_enabled = USE_CACHE and cache is not None
    if cache_enabled and nif in cache:
        return cache[nif]

    @retry(
        reraise=True,
        stop=stop_after_attempt(3),
        wait=wait_exponential(multiplier=0.5, min=0.5, max=2),
        retry=retry_if_exception_type((TransientAPIError, requests.Timeout, requests.ConnectionError)),
    )
    def _do_request() -> dict[str, Any] | None:
        _throttle_if_needed()
        url = _build_url(nif)
        response = session.get(url, timeout=API_TIMEOUT)
        if 500 <= response.status_code < 600:
            raise TransientAPIError(f"Erro {response.status_code} ao obter dados do contribuinte")
        if response.status_code >= 400:
            return None
        try:
            payload = response.json()
        except json.JSONDecodeError:
            return None
        if not isinstance(payload, Mapping):
            return None
        success = payload.get("success")
        if not success:
            return None
        data = payload.get("data")
        if not isinstance(data, Mapping):
            return {}
        return {
            "companyName": data.get("companyName"),
            "gsmc": data.get("gsmc"),
            "nsrdz": data.get("nsrdz"),
            "hdzt": data.get("hdzt"),
        }

    try:
        result = _do_request()
    except RetryError:
        result = None
    if cache_enabled:
        cache[nif] = result
    return result


def aplicar_regras(
    linha: dict[str, Any],
    api: dict[str, Any] | None,
    nif_norm: str,
    primeiro_codigo_por_nif: Mapping[str, Any],
    classificacao_nif: str | None = None,
    nif_portugal: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Aplica as regras de atualização às colunas principais."""
    resultado = dict(linha)
    codigo = resultado.get("Codigo")

    if classificacao_nif is None:
        classificacao_nif = classificar_nif_ao(linha.get("NIF"))

    if classificacao_nif != "possivelmente_correto" or nif_portugal is not None:
        if nif_portugal is not None:
            pais = nif_portugal.get("pais")
            if pais:
                resultado["Pais"] = pais
            nome_pt = nif_portugal.get("nome")
            if nome_pt:
                resultado["Nome"] = nome_pt
            morada_pt = nif_portugal.get("morada")
            if morada_pt:
                resultado["Morada"] = morada_pt
            localidade_pt = nif_portugal.get("localidade")
            mensagem_pt = nif_portugal.get("mensagem")
            if localidade_pt:
                resultado["Localidade"] = localidade_pt
            elif mensagem_pt:
                resultado["Localidade"] = mensagem_pt
            return resultado

        resultado["Localidade"] = _mensagem_nif_invalido(classificacao_nif)
        return resultado

    mensagem_invalido: str | None = None

    if api is None:
        mensagem_invalido = _mensagem_nif_invalido("possivelmente_errado")
    else:
        company = (api.get("companyName") or "").strip()
        gsmc = (api.get("gsmc") or "").strip()
        morada = (api.get("nsrdz") or "").strip()
        estado = api.get("hdzt")

        if company:
            resultado["Nome"] = company
        elif gsmc:
            resultado["Nome"] = gsmc

        if morada:
            resultado["Morada"] = morada

        if not isinstance(estado, str) or estado.upper() != "ACTIVE":
            resultado["Localidade"] = "Contribuinte INACTIVO na AGT"

    primeiro_codigo = primeiro_codigo_por_nif.get(nif_norm)
    duplicado = primeiro_codigo is not None and primeiro_codigo != codigo

    if duplicado:
        resultado["Localidade"] = f"NIF DUPLICADO - {primeiro_codigo}"
    elif mensagem_invalido is not None:
        resultado["Localidade"] = mensagem_invalido

    return resultado


def _normalize_header_key(value: str) -> str:
    """Normaliza nomes de colunas removendo acentos e caracteres especiais."""

    normalized = unicodedata.normalize("NFKD", str(value))
    normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
    normalized = normalized.lower()
    normalized = re.sub(r"[^a-z0-9]+", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized


CANONICAL_COLUMNS: dict[str, set[str]] = {
    "Codigo": {
        "codigo",
        "cod",
        "cod_cliente",
        "codigo_cliente",
        "codigo_de_cliente",
        "client_code",
    },
    "NIF": {
        "nif",
        "nif_cliente",
        "numero_contribuinte",
        "num_contribuinte",
        "contribuinte",
        "num_contribuinte",
        "n_contribuinte",
        "no_contribuinte",
        "nro_contribuinte",
        "num_contrib",
        "numero_nif",
        "nif_numero",
    },
    "Nome": {
        "nome",
        "nome_cliente",
        "cliente",
        "designacao",
        "designacao_social",
        "razao_social",
    },
    "Morada": {
        "morada",
        "endereco",
        "endereco_cliente",
        "endereco_postal",
        "endereco_fiscal",
        "address",
    },
    "Localidade": {
        "localidade",
        "cidade",
        "municipio",
        "localizacao",
        "local",
        "cidade_cliente",
        "city",
    },
    "Pais": {
        "pais",
        "país",
        "country",
        "pais_cliente",
        "país_cliente",
    },
}

OPTIONAL_CANONICAL_COLUMNS = {"Pais"}


def _mapear_colunas(columns: Iterable[str]) -> dict[str, str]:
    """Devolve um mapping de colunas canónicas para os nomes reais no ficheiro."""

    normalized_columns: dict[str, str] = {}
    for col in columns:
        key = _normalize_header_key(col)
        if key and key not in normalized_columns:
            normalized_columns[key] = col

    canonical_to_actual: dict[str, str] = {}
    for canonical, synonyms in CANONICAL_COLUMNS.items():
        for candidate in synonyms:
            candidate_key = _normalize_header_key(candidate)
            if candidate_key in normalized_columns:
                canonical_to_actual[canonical] = normalized_columns[candidate_key]
                break
    return canonical_to_actual


def corrigir_excel(input_path: str, output_path: str | None = None) -> str:
    """Corrige um ficheiro Excel de clientes usando dados da AGT."""
    global LAST_SUMMARY
    df = pd.read_excel(input_path, dtype=object)
    columns = list(df.columns)
    canonical_to_actual = _mapear_colunas(columns)
    required = set(CANONICAL_COLUMNS) - OPTIONAL_CANONICAL_COLUMNS
    missing = [col for col in required if col not in canonical_to_actual]
    if missing:
        raise ValueError(f"Colunas obrigatórias em falta: {', '.join(missing)}")

    records = df.to_dict("records")

    primeiro_codigo_por_nif: dict[str, Any] = {}
    contagem_por_nif: dict[str, int] = {}
    for row in records:
        nif_norm = normalizar_nif(row.get(canonical_to_actual["NIF"]))
        if nif_norm:
            contagem_por_nif[nif_norm] = contagem_por_nif.get(nif_norm, 0) + 1
            if nif_norm not in primeiro_codigo_por_nif:
                primeiro_codigo_por_nif[nif_norm] = row.get(canonical_to_actual["Codigo"])

    nifs_com_duplicados = {nif for nif, count in contagem_por_nif.items() if count > 1}
    duplicados_marcados = 0
    invalidos = 0
    validos = 0

    cache: dict[str, dict[str, Any] | None] | None = {} if USE_CACHE else None
    cache_pt: dict[str, dict[str, Any] | None] | None = {} if USE_CACHE else None
    session = requests.Session()
    estilos_linhas: list[str] = []
    novos_registos: list[dict[str, Any]] = []

    try:
        for row in records:
            novo = dict(row)
            canonical_linha = {
                chave: novo.get(actual)
                for chave, actual in canonical_to_actual.items()
            }
            classificacao_nif = classificar_nif_ao(canonical_linha["NIF"])
            nif_norm = normalizar_nif(canonical_linha["NIF"])

            api_data = None
            if classificacao_nif == "possivelmente_correto" and nif_norm:
                api_data = fetch_taxpayer(nif_norm, session, cache)

            nif_portugal = None
            if nif_norm and (
                classificacao_nif != "possivelmente_correto"
                or (classificacao_nif == "possivelmente_correto" and not api_data)
            ):
                nif_portugal = avaliar_nif_portugues(nif_norm, session, cache_pt)

            resultado = aplicar_regras(
                canonical_linha,
                api_data,
                nif_norm,
                primeiro_codigo_por_nif,
                classificacao_nif=classificacao_nif,
                nif_portugal=nif_portugal,
            )

            localidade_resultado = resultado.get("Localidade")
            if isinstance(localidade_resultado, str) and localidade_resultado.startswith("NIF INVALIDO"):
                invalidos += 1
            else:
                validos += 1

            if nif_norm and primeiro_codigo_por_nif.get(nif_norm) != resultado.get("Codigo"):
                duplicados_marcados += 1

            for chave, actual in canonical_to_actual.items():
                novo[actual] = resultado[chave]
            novos_registos.append(novo)

            estado_linha = "normal"
            if isinstance(localidade_resultado, str):
                if localidade_resultado.startswith("NIF INVALIDO"):
                    estado_linha = "nif_invalido"
                elif localidade_resultado.startswith("NIF DUPLICADO -"):
                    estado_linha = "nif_duplicado"
                elif localidade_resultado == "Contribuinte INACTIVO na AGT":
                    estado_linha = "contribuinte_inactivo"

            if estado_linha == "normal":
                def _norm(value: Any) -> str:
                    if pd.isna(value):
                        return ""
                    return str(value)

                campos_a_comparar = ("Nome", "Morada")
                alterado = any(
                    _norm(resultado.get(campo)) != _norm(canonical_linha.get(campo))
                    for campo in campos_a_comparar
                )
                if alterado:
                    estado_linha = "actualizado"

            estilos_linhas.append(estado_linha)
    finally:
        session.close()

    novo_df = pd.DataFrame(novos_registos, columns=columns)

    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_corrigido{ext or '.xlsx'}"

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        novo_df.to_excel(writer, index=False)

        folha = next(iter(writer.sheets.values()))
        max_col = folha.max_column

        font_verde_negrito = Font(color="FF006100", bold=True)
        font_vermelho = Font(color="FFFF0000")
        font_branco = Font(color="FFFFFFFF")
        fill_laranja = PatternFill(fill_type="solid", start_color="FFFFA500", end_color="FFFFA500")
        fill_vermelho = PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")

        for idx, estado in enumerate(estilos_linhas, start=2):
            linha = next(folha.iter_rows(min_row=idx, max_row=idx, min_col=1, max_col=max_col))
            if estado == "actualizado":
                for cell in linha:
                    cell.font = font_verde_negrito
            elif estado == "nif_invalido":
                for cell in linha:
                    cell.font = font_vermelho
            elif estado == "nif_duplicado":
                for cell in linha:
                    cell.font = font_branco
                    cell.fill = fill_laranja
            elif estado == "contribuinte_inactivo":
                for cell in linha:
                    cell.font = font_branco
                    cell.fill = fill_vermelho

    LAST_SUMMARY = {
        "linhas": len(novos_registos),
        "validos": validos,
        "invalidos": invalidos,
        "nifs_duplicados": len(nifs_com_duplicados),
        "duplicados_marcados": duplicados_marcados,
    }

    return output_path
