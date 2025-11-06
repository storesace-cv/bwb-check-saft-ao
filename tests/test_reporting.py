from __future__ import annotations

from decimal import Decimal

from openpyxl import load_workbook

from saftao.schema import load_audit_file
from saftao.utils.reporting import aggregate_documents, write_excel_report


SAMPLE_XML = """
<AuditFile xmlns="urn:OECD:StandardAuditFile-Tax:AO_1.01_01">
  <Header>
    <CompanyID>123456789</CompanyID>
  </Header>
  <SourceDocuments>
    <SalesInvoices>
      <NumberOfEntries>2</NumberOfEntries>
      <TotalDebit>0.00</TotalDebit>
      <TotalCredit>0.00</TotalCredit>
      <Invoice>
        <InvoiceNo>FT 1</InvoiceNo>
        <InvoiceType>FT</InvoiceType>
        <InvoiceDate>2023-01-01</InvoiceDate>
        <CustomerID>C1</CustomerID>
        <DocumentTotals>
          <TaxPayable>20.00</TaxPayable>
          <NetTotal>100.00</NetTotal>
          <GrossTotal>120.00</GrossTotal>
        </DocumentTotals>
      </Invoice>
      <Invoice>
        <InvoiceNo>NC 1</InvoiceNo>
        <InvoiceType>NC</InvoiceType>
        <InvoiceDate>2023-01-02</InvoiceDate>
        <CustomerID>C1</CustomerID>
        <DocumentTotals>
          <TaxPayable>-3.00</TaxPayable>
          <NetTotal>-15.00</NetTotal>
          <GrossTotal>-18.00</GrossTotal>
        </DocumentTotals>
      </Invoice>
    </SalesInvoices>
    <WorkingDocuments>
      <NumberOfEntries>2</NumberOfEntries>
      <TotalDebit>0.00</TotalDebit>
      <TotalCredit>0.00</TotalCredit>
      <WorkDocument>
        <DocumentNumber>GT 1</DocumentNumber>
        <DocumentType>GT</DocumentType>
        <WorkDate>2023-01-03</WorkDate>
        <CustomerID>C2</CustomerID>
        <DocumentTotals>
          <TaxPayable>2.50</TaxPayable>
          <NetTotal>12.50</NetTotal>
          <GrossTotal>15.00</GrossTotal>
        </DocumentTotals>
      </WorkDocument>
      <WorkDocument>
        <DocumentNumber>RQ 1</DocumentNumber>
        <DocumentType>RQ</DocumentType>
        <WorkDate>2023-01-04</WorkDate>
        <CustomerID>C3</CustomerID>
        <DocumentTotals>
          <TaxPayable>1.00</TaxPayable>
          <NetTotal>5.00</NetTotal>
          <GrossTotal>6.00</GrossTotal>
        </DocumentTotals>
      </WorkDocument>
    </WorkingDocuments>
  </SourceDocuments>
</AuditFile>
""".strip()


def test_aggregate_documents_totals(tmp_path):
    xml_path = tmp_path / "sample.xml"
    xml_path.write_text(SAMPLE_XML, encoding="utf-8")

    _tree, root, namespace = load_audit_file(xml_path)
    report_data = aggregate_documents(root, namespace)

    assert set(report_data.totals_by_type) == {"FT", "NC"}
    assert report_data.totals_by_type["FT"].net_total == Decimal("100.00")
    assert report_data.totals_by_type["FT"].tax_total == Decimal("20.00")
    assert report_data.totals_by_type["FT"].gross_total == Decimal("120.00")
    assert report_data.totals_by_type["NC"].net_total == Decimal("-15.00")
    assert report_data.totals_by_type["NC"].tax_total == Decimal("-3.00")
    assert report_data.totals_by_type["NC"].gross_total == Decimal("-18.00")

    assert report_data.overall_totals.net_total == Decimal("85.00")
    assert report_data.overall_totals.tax_total == Decimal("17.00")
    assert report_data.overall_totals.gross_total == Decimal("102.00")

    assert len(report_data.non_accounting_documents) == 2
    first_doc = report_data.non_accounting_documents[0]
    assert first_doc.document_type == "GT"
    assert first_doc.document_number == "GT 1"
    assert first_doc.totals.net_total == Decimal("12.50")
    assert first_doc.totals.tax_total == Decimal("2.50")
    assert first_doc.totals.gross_total == Decimal("15.00")


def test_excel_report_contains_non_accounting_section(tmp_path):
    xml_path = tmp_path / "sample.xml"
    xml_path.write_text(SAMPLE_XML, encoding="utf-8")
    excel_path = tmp_path / "report.xlsx"

    _tree, root, namespace = load_audit_file(xml_path)
    report_data = aggregate_documents(root, namespace)
    write_excel_report(report_data, excel_path)

    workbook = load_workbook(excel_path)

    summary_rows = list(workbook["Resumo"].iter_rows(values_only=True))
    assert summary_rows[0] == ("Tipo", "Total sem IVA", "IVA", "Total com IVA")
    ft_row = summary_rows[1]
    assert ft_row[0] == "FT"
    assert Decimal(str(ft_row[1])) == Decimal("100")
    assert Decimal(str(ft_row[2])) == Decimal("20")
    assert Decimal(str(ft_row[3])) == Decimal("120")

    other_rows = list(
        workbook["Documentos não contabilísticos"].iter_rows(values_only=True)
    )
    assert other_rows[0] == (
        "Tipo",
        "Número",
        "Data",
        "Cliente",
        "Total sem IVA",
        "IVA",
        "Total com IVA",
    )
    assert ("GT", "GT 1") == other_rows[1][:2]
    assert ("RQ", "RQ 1") == other_rows[2][:2]
