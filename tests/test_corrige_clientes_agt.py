from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
import pytest
from openpyxl import load_workbook

import tools.corrige_clientes_agt as corrige

from tools.corrige_clientes_agt import (
    aplicar_regras,
    classificar_nif_ao,
    corrigir_excel,
    fetch_taxpayer,
    normalizar_nif,
    set_fetch_settings,
)


class DummyResponse:
    def __init__(self, status_code: int, payload: Any | None = None):
        self.status_code = status_code
        self._payload = payload

    def json(self):
        if isinstance(self._payload, Exception):
            raise self._payload
        return self._payload


class DummySession:
    def __init__(self, responses: dict[str, DummyResponse]):
        self.responses = responses
        self.calls: list[str] = []

    def get(self, url: str, timeout: float):
        self.calls.append(url)
        nif = url.rsplit("/", 1)[-1]
        return self.responses.get(nif, DummyResponse(404, {"success": False}))

    def close(self):
        pass


def test_normalizar_nif_remove_nao_alfanumericos():
    assert normalizar_nif("001.234.567-8") == "0012345678"
    assert normalizar_nif("abc-123") == "ABC123"


@pytest.mark.parametrize(
    "valor, esperado",
    [
        ("", "manifestamente_errado"),
        (None, "manifestamente_errado"),
        ("999999999", "possivelmente_correto"),
        ("5000000000", "possivelmente_correto"),
        ("003489072LA037", "possivelmente_correto"),
        ("12345", "manifestamente_errado"),
        ("ABCDE12345", "manifestamente_errado"),
        ("1234567", "possivelmente_errado"),
    ],
)
def test_classificar_nif_ao(valor: Any, esperado: str) -> None:
    assert classificar_nif_ao(valor) == esperado


def test_aplicar_regras_preferencia_company_name():
    linha = {"Codigo": "1", "NIF": "5000000000", "Nome": "Original", "Morada": "Rua", "Localidade": "Cidade"}
    api = {"companyName": "Empresa", "gsmc": "Oficial", "nsrdz": "Endereco", "hdzt": "ACTIVE"}
    resultado = aplicar_regras(linha, api, "5000000000", {"5000000000": "1"})
    assert resultado["Nome"] == "Empresa"
    assert resultado["Morada"] == "Endereco"
    assert resultado["Localidade"] == "Cidade"


def test_aplicar_regras_inactivo_altera_localidade():
    linha = {"Codigo": "1", "NIF": "5000000000", "Nome": "Original", "Morada": "Rua", "Localidade": "Cidade"}
    api = {"gsmc": "Oficial", "nsrdz": "Endereco", "hdzt": "SUSPENDED"}
    resultado = aplicar_regras(linha, api, "5000000000", {"5000000000": "1"})
    assert resultado["Localidade"] == "Contribuinte INACTIVO na AGT"


def test_aplicar_regras_nif_invalido():
    linha = {"Codigo": "1", "NIF": "", "Nome": "Original", "Morada": "Rua", "Localidade": "Cidade"}
    resultado = aplicar_regras(linha, None, "", {}, classificacao_nif="manifestamente_errado")
    assert resultado["Localidade"] == "NIF INVALIDO | Manifestamente errado"


def test_aplicar_regras_nif_duplicado():
    linha = {"Codigo": "2", "NIF": "5000000000", "Nome": "Original", "Morada": "Rua", "Localidade": "Cidade"}
    resultado = aplicar_regras(linha, {}, "5000000000", {"5000000000": "1"})
    assert resultado["Localidade"] == "NIF DUPLICADO - 1"


def test_fetch_taxpayer_usa_cache(monkeypatch):
    responses = {"500": DummyResponse(200, {"success": True, "data": {"companyName": "Empresa"}})}
    session = DummySession(responses)
    cache: dict[str, Any] = {}
    set_fetch_settings(rate_limit=0, timeout=10, use_cache=True)

    # primeira chamada deve realizar request
    resultado1 = fetch_taxpayer("500", session, cache)
    assert resultado1["companyName"] == "Empresa"
    assert len(session.calls) == 1

    # segunda chamada usa cache
    resultado2 = fetch_taxpayer("500", session, cache)
    assert resultado2["companyName"] == "Empresa"
    assert len(session.calls) == 1


def test_corrigir_excel_integration(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    set_fetch_settings(rate_limit=0, timeout=10, use_cache=False)

    dados = pd.DataFrame(
        [
            {"Codigo": "C1", "NIF": "500000000", "Nome": "Orig1", "Morada": "Rua 1", "Localidade": "Cidade"},
            {"Codigo": "C2", "NIF": "500000000", "Nome": "Orig2", "Morada": "Rua 2", "Localidade": "Cidade"},
            {"Codigo": "C3", "NIF": "", "Nome": "Orig3", "Morada": "Rua 3", "Localidade": "Cidade"},
        ]
    )
    input_path = tmp_path / "clientes.xlsx"
    dados.to_excel(input_path, index=False)

    responses = {
        "500000000": DummyResponse(
            200,
            {
                "success": True,
                "data": {
                    "companyName": "Empresa Nova",
                    "gsmc": "Nome Oficial",
                    "nsrdz": "Endereco Atualizado",
                    "hdzt": "ACTIVE",
                },
            },
        )
    }

    monkeypatch.setattr(
        "tools.corrige_clientes_agt.requests.Session",
        lambda: DummySession(responses),
    )

    output_path = corrigir_excel(str(input_path))
    assert Path(output_path).exists()

    result = pd.read_excel(output_path)
    assert list(result["Nome"]) == ["Empresa Nova", "Empresa Nova", "Orig3"]
    assert list(result["Morada"]) == ["Endereco Atualizado", "Endereco Atualizado", "Rua 3"]
    assert result.loc[0, "Localidade"] == "Cidade"
    assert result.loc[1, "Localidade"] == "NIF DUPLICADO - C1"
    assert result.loc[2, "Localidade"] == "NIF INVALIDO | Manifestamente errado"

    summary = corrige.LAST_SUMMARY
    assert summary is not None
    assert summary["linhas"] == 3
    assert summary["invalidos"] == 1
    assert summary["nifs_duplicados"] == 1
    assert summary["duplicados_marcados"] == 1


def test_corrigir_excel_custom_output_path(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    set_fetch_settings(rate_limit=0, timeout=10, use_cache=False)

    dados = pd.DataFrame(
        [
            {"Codigo": "C1", "NIF": "500000000", "Nome": "Orig", "Morada": "Rua", "Localidade": "Cidade"},
        ]
    )
    input_path = tmp_path / "clientes.xlsx"
    dados.to_excel(input_path, index=False)

    responses = {
        "500000000": DummyResponse(
            200,
            {
                "success": True,
                "data": {
                    "companyName": "Empresa Nova",
                    "gsmc": "Nome Oficial",
                    "nsrdz": "Endereco Atualizado",
                    "hdzt": "ACTIVE",
                },
            },
        )
    }

    monkeypatch.setattr(
        "tools.corrige_clientes_agt.requests.Session",
        lambda: DummySession(responses),
    )

    destino = tmp_path / "destino" / "clientes"
    destino.mkdir(parents=True)
    output_path = destino / "clientes_corrigido.xlsx"

    result_path = corrigir_excel(str(input_path), str(output_path))

    assert Path(result_path) == output_path
    assert output_path.exists()


def test_corrigir_excel_aplica_formatacao(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    set_fetch_settings(rate_limit=0, timeout=10, use_cache=False)

    dados = pd.DataFrame(
        [
            {"Codigo": "C1", "NIF": "500000000", "Nome": "Orig1", "Morada": "Rua 1", "Localidade": "Cidade"},
            {"Codigo": "C2", "NIF": "500000000", "Nome": "Orig2", "Morada": "Rua 2", "Localidade": "Cidade"},
            {"Codigo": "C3", "NIF": "", "Nome": "Orig3", "Morada": "Rua 3", "Localidade": "Cidade"},
            {"Codigo": "C4", "NIF": "600000000", "Nome": "Orig4", "Morada": "Rua 4", "Localidade": "Cidade"},
        ]
    )
    input_path = tmp_path / "clientes.xlsx"
    dados.to_excel(input_path, index=False)

    responses = {
        "500000000": DummyResponse(
            200,
            {
                "success": True,
                "data": {
                    "companyName": "Empresa Nova",
                    "gsmc": "Nome Oficial",
                    "nsrdz": "Endereco Atualizado",
                    "hdzt": "ACTIVE",
                },
            },
        ),
        "600000000": DummyResponse(
            200,
            {
                "success": True,
                "data": {
                    "companyName": "Empresa Inactiva",
                    "gsmc": "Empresa Inactiva",
                    "nsrdz": "Endereco Inactivo",
                    "hdzt": "SUSPENDED",
                },
            },
        ),
    }

    monkeypatch.setattr(
        "tools.corrige_clientes_agt.requests.Session",
        lambda: DummySession(responses),
    )

    output_path = corrigir_excel(str(input_path))
    wb = load_workbook(output_path)
    try:
        ws = wb.active

        def font_rgb(cell):
            return None if cell.font.color is None else cell.font.color.rgb

        # Registo actualizado (linha 2)
        cell = ws.cell(row=2, column=1)
        assert font_rgb(cell) == "FF006100"
        assert cell.font.bold is True
        assert ws.cell(row=2, column=1).fill.patternType in (None, "none")

        # NIF duplicado (linha 3)
        cell = ws.cell(row=3, column=1)
        assert font_rgb(cell) == "FFFFFFFF"
        assert cell.fill.patternType == "solid"
        assert cell.fill.fgColor.rgb == "FFFFA500"

        # NIF inválido (linha 4)
        cell = ws.cell(row=4, column=1)
        assert font_rgb(cell) == "FFFF0000"
        assert cell.fill.patternType in (None, "none")

        # Contribuinte inactivo (linha 5)
        cell = ws.cell(row=5, column=1)
        assert font_rgb(cell) == "FFFFFFFF"
        assert cell.fill.patternType == "solid"
        assert cell.fill.fgColor.rgb == "FFFF0000"
    finally:
        wb.close()


def test_corrigir_excel_accepts_synonym_headers(tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
    set_fetch_settings(rate_limit=0, timeout=10, use_cache=False)

    dados = pd.DataFrame(
        [
            {
                "Cod Cliente": "C1",
                "Nº Contribuinte": "500000000",
                "Designação": "Orig",
                "Endereço Fiscal": "Rua",
                "Cidade": "Cidade",
            }
        ]
    )
    input_path = tmp_path / "clientes.xlsx"
    dados.to_excel(input_path, index=False)

    responses = {
        "500000000": DummyResponse(
            200,
            {
                "success": True,
                "data": {
                    "companyName": "Empresa Nova",
                    "gsmc": "Nome Oficial",
                    "nsrdz": "Endereco Atualizado",
                    "hdzt": "ACTIVE",
                },
            },
        )
    }

    monkeypatch.setattr(
        "tools.corrige_clientes_agt.requests.Session",
        lambda: DummySession(responses),
    )

    output_path = corrigir_excel(str(input_path))
    result = pd.read_excel(output_path)

    assert "Cod Cliente" in result.columns
    assert result.loc[0, "Designação"] == "Empresa Nova"
    assert result.loc[0, "Endereço Fiscal"] == "Endereco Atualizado"
