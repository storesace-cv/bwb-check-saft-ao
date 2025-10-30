"""Soft auto-fix routines for SAFT AO XML files."""

from __future__ import annotations

import functools
import os
import unicodedata
from dataclasses import dataclass
from math import ceil
from pathlib import Path
from typing import Iterable, Sequence, TYPE_CHECKING

if TYPE_CHECKING:
    import tkinter as tk

from lxml import etree

from ..logging import ExcelLogger, ExcelLoggerConfig
from ..rules import (
    collect_invoice_customer_ids,
    collect_masterfile_customer_ids,
)
from ..utils import detect_namespace
from ..validator import ValidationIssue

_EXCEL_ENV_VARIABLE = "BWB_SAFTAO_CUSTOMER_FILE"
_REPO_ROOT = Path(__file__).resolve().parents[3]
_DEFAULT_ADDONS_DIR = _REPO_ROOT / "work" / "origem" / "addons"
_DEFAULT_CUSTOMER_FILENAME = "Listagem_de_Clientes.xlsx"
_COUNTRY_CODES_PATH = _REPO_ROOT / "docs" / "paises_iso_alpha2_pt.md"


@dataclass
class _CustomerRecord:
    """Representa os dados mínimos necessários para criar um cliente."""

    customer_id: str
    company_name: str
    tax_id: str
    city: str
    country: str
    telephone: str
    source_path: Path


def apply_soft_fixes(path: Path) -> Iterable[ValidationIssue]:
    """Apply non-destructive corrections to the given file.

    The intent is to migrate the behaviours from ``saft_ao_autofix_soft.py``
    into this module, exposing a clean function that yields issues which were
    auto-resolved.  The stub raises :class:`NotImplementedError` until that
    work happens.
    """

    raise NotImplementedError("Soft auto-fixes still need to be implemented")


def normalize_invoice_type_vd(path: Path) -> Iterable[ValidationIssue]:
    """Replace ``InvoiceType="VD"`` entries with ``"FR"`` in-place."""

    tree = etree.parse(str(path))
    issues = normalize_invoice_type_vd_tree(tree)
    if issues:
        tree.write(
            str(path),
            encoding="utf-8",
            xml_declaration=True,
            pretty_print=True,
        )
    return issues


def normalize_invoice_type_vd_tree(
    tree: etree._ElementTree,
) -> list[ValidationIssue]:
    """Apply the ``VD`` → ``FR`` normalisation to an in-memory XML tree."""

    root = tree.getroot()
    ns_uri = detect_namespace(root)
    namespaces = {"n": ns_uri} if ns_uri else None

    if namespaces:
        xpath = ".//n:SourceDocuments/n:SalesInvoices/n:Invoice"
        invoices = root.xpath(xpath, namespaces=namespaces)
    else:
        invoices = root.findall(".//SourceDocuments/SalesInvoices/Invoice")

    issues: list[ValidationIssue] = []
    for invoice in invoices:
        invoice_type = _find_child(invoice, "InvoiceType", ns_uri)
        if invoice_type is None:
            continue
        value = (invoice_type.text or "").strip()
        if value.upper() != "VD":
            continue

        invoice_no = _find_child_text(invoice, "InvoiceNo", ns_uri) or "(sem número)"
        invoice_type.text = "FR"
        issues.append(
            ValidationIssue(
                f"Invoice '{invoice_no}': InvoiceType 'VD' substituído por 'FR'.",
                code="FIX_INVOICE_TYPE",
                details={"invoice": invoice_no},
            )
        )

    return issues


def ensure_invoice_customers_exported(path: Path) -> Iterable[ValidationIssue]:
    """Ensure every customer referenced by an invoice is present in MasterFiles.

    Esta rotina detecta clientes utilizados nas facturas que ainda não estejam
    presentes no bloco ``MasterFiles/Customer`` do ficheiro SAF-T. Caso sejam
    identificadas ausências, o utilizador é solicitado a indicar o ficheiro
    Excel com a tabela de clientes (padrão ``work/origem/addons``). Os registos
    em falta são adicionados automaticamente ao XML.
    """

    tree = etree.parse(str(path))
    issues = ensure_invoice_customers_exported_tree(tree)
    if issues:
        tree.write(
            str(path),
            encoding="utf-8",
            xml_declaration=True,
            pretty_print=True,
        )
    return issues


def ensure_invoice_customers_exported_tree(
    tree: etree._ElementTree,
) -> list[ValidationIssue]:
    """Ensure every invoice customer exists in ``MasterFiles`` for ``tree``."""

    root = tree.getroot()
    ns_uri = detect_namespace(root)

    invoice_ids = collect_invoice_customer_ids(root, ns_uri)
    existing_ids = collect_masterfile_customer_ids(root, ns_uri)

    missing_ids = [cid for cid in invoice_ids if cid not in existing_ids]
    if not missing_ids:
        return []

    records = _gather_customer_records(missing_ids)

    masterfiles = _ensure_masterfiles_node(root, ns_uri)

    issues: list[ValidationIssue] = []
    for customer_id in missing_ids:
        record = records[customer_id]
        _append_customer(masterfiles, ns_uri, record)
        issues.append(
            ValidationIssue(
                f"Cliente '{customer_id}' adicionado ao MasterFiles com dados de "
                f"'{record.source_path.name}'.",
                code="AUTOADD_CUSTOMER",
                details={
                    "customer_id": customer_id,
                    "source": str(record.source_path),
                },
            )
        )

    return issues


def log_soft_fixes(issues: Iterable[ValidationIssue], *, destination: Path) -> None:
    """Persist the soft fixes to a spreadsheet log."""

    logger = ExcelLogger(
        ExcelLoggerConfig(columns=("code", "message"), filename=str(destination))
    )
    logger.write_rows(issues)


def _gather_customer_records(missing_ids: list[str]) -> dict[str, _CustomerRecord]:
    env_path = os.environ.get(_EXCEL_ENV_VARIABLE)
    if env_path:
        excel_path = Path(env_path).expanduser()
        if not excel_path.exists():
            raise FileNotFoundError(
                f"O ficheiro Excel definido em {_EXCEL_ENV_VARIABLE} não existe: {excel_path}"
            )
        return _map_records_for_missing_ids(excel_path, missing_ids)

    default_excel = _DEFAULT_ADDONS_DIR / _DEFAULT_CUSTOMER_FILENAME
    if default_excel.exists():
        return _map_records_for_missing_ids(default_excel, missing_ids)

    return _gather_records_interactively(missing_ids)


def _map_records_for_missing_ids(
    excel_path: Path, missing_ids: list[str]
) -> dict[str, _CustomerRecord]:
    records = _load_records_from_excel(excel_path)
    result: dict[str, _CustomerRecord] = {}
    missing_from_file: list[str] = []
    for customer_id in missing_ids:
        data = records.get(customer_id)
        if data is None:
            missing_from_file.append(customer_id)
            continue
        result[customer_id] = _CustomerRecord(**data, source_path=excel_path)
    if missing_from_file:
        missing_str = ", ".join(missing_from_file)
        raise ValueError(
            "Os seguintes clientes em falta não foram encontrados no ficheiro "
            f"{excel_path}: {missing_str}"
        )
    return result


def _gather_records_interactively(
    missing_ids: list[str],
) -> dict[str, _CustomerRecord]:
    import tkinter as tk

    root = tk.Tk()
    root.withdraw()

    _DEFAULT_ADDONS_DIR.mkdir(parents=True, exist_ok=True)

    try:
        _show_missing_customers_dialog(root, missing_ids)

        excel_path = _DEFAULT_ADDONS_DIR / _DEFAULT_CUSTOMER_FILENAME

        if not excel_path.exists():
            _show_message(
                "error",
                "Ficheiro de clientes não encontrado",
                (
                    "Não foi possível localizar o ficheiro fixo de clientes no "
                    f"directório {_DEFAULT_ADDONS_DIR}."
                ),
                parent=root,
            )
            excel_path = _prompt_for_customer_file(root, initialdir=_DEFAULT_ADDONS_DIR)
            if excel_path is None:
                raise FileNotFoundError(
                    "Operação cancelada: ficheiro de clientes obrigatório não seleccionado."
                )

        try:
            collected = _map_records_for_missing_ids(excel_path, missing_ids)
        except ValueError as exc:
            _show_message(
                "error",
                "Clientes em falta no ficheiro",
                str(exc),
                parent=root,
            )
            raise
        except Exception as exc:  # pragma: no cover - interface interativa
            _show_message(
                "error",
                "Erro ao ler ficheiro Excel",
                str(exc),
                parent=root,
            )
            raise

        return collected
    finally:
        root.destroy()


def _show_message(
    kind: str,
    title: str,
    text: str,
    informative_text: str | None = None,
    *,
    parent: tk.Misc | None = None,
) -> None:
    from tkinter import messagebox

    message = text if informative_text is None else f"{text}\n\n{informative_text}"
    options = {"parent": parent} if parent is not None else {}
    if kind == "error":
        messagebox.showerror(title, message, **options)
    elif kind == "warning":
        messagebox.showwarning(title, message, **options)
    else:
        messagebox.showinfo(title, message, **options)


def _show_missing_customers_dialog(parent: "tk.Misc", missing_ids: Sequence[str]) -> None:
    import tkinter as tk
    from tkinter import ttk

    window = tk.Toplevel(parent)
    window.title("Clientes em falta no MasterFiles")
    window.transient(parent)
    window.grab_set()
    window.resizable(True, True)
    window.geometry("800x400")
    window.minsize(800, 400)

    background_color = window.cget("background")

    description = (
        "Foram detectados clientes nas facturas que não existem no MasterFiles. "
        "Indique o ficheiro de clientes que deve ser utilizado para completar os registos."
    )

    label = tk.Label(
        window,
        text=description,
        wraplength=760,
        justify="left",
        anchor="w",
        background=background_color,
        foreground="#000000",
    )
    label.pack(fill="x", padx=16, pady=(16, 8))

    frame = ttk.Frame(window, padding=0)
    frame.pack(fill="both", expand=True, padx=16, pady=(0, 8))

    canvas = tk.Canvas(
        frame,
        borderwidth=0,
        highlightthickness=0,
        background=background_color,
    )
    canvas.pack(side="left", fill="both", expand=True)

    scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)

    canvas.configure(yscrollcommand=scrollbar.set)

    content = tk.Frame(canvas, background=background_color, borderwidth=0, highlightthickness=0)
    window_item = canvas.create_window((0, 0), window=content, anchor="nw")

    columns = max(1, min(4, len(missing_ids)))
    for index, customer_id in enumerate(missing_ids):
        row = index // columns
        column = index % columns
        label = tk.Label(
            content,
            text=customer_id,
            anchor="center",
            width=18,
            background=background_color,
            foreground="#000000",
            padx=8,
            pady=6,
        )
        label.grid(row=row, column=column, sticky="nsew", padx=4, pady=4)

    for column in range(columns):
        content.columnconfigure(column, weight=1)

    needs_scrollbar = {"visible": False}

    def _update_scrollbar_visibility() -> None:
        content_height = content.winfo_reqheight()
        canvas_height = canvas.winfo_height()
        if content_height > canvas_height:
            if not needs_scrollbar["visible"]:
                scrollbar.pack(side="right", fill="y")
                needs_scrollbar["visible"] = True
        else:
            if needs_scrollbar["visible"]:
                scrollbar.pack_forget()
                needs_scrollbar["visible"] = False

    def _on_content_configure(event: tk.Event) -> None:  # pragma: no cover - UI callback
        canvas.configure(scrollregion=canvas.bbox("all"))
        _update_scrollbar_visibility()

    def _on_canvas_configure(event: tk.Event) -> None:  # pragma: no cover - UI callback
        canvas.itemconfigure(window_item, width=event.width)
        _update_scrollbar_visibility()

    content.bind("<Configure>", _on_content_configure)
    canvas.bind("<Configure>", _on_canvas_configure)

    button = ttk.Button(window, text="Continuar", command=window.destroy)
    button.pack(pady=(0, 16))

    window.update_idletasks()
    _update_scrollbar_visibility()
    window.wait_window()


def _prompt_for_customer_file(
    parent: "tk.Misc", *, initialdir: Path | None = None
) -> Path | None:
    from tkinter import filedialog

    selected = filedialog.askopenfilename(
        parent=parent,
        title="Seleccione o ficheiro de clientes",
        initialdir=str(initialdir) if initialdir else None,
        filetypes=[
            ("Ficheiros Excel", "*.xlsx"),
            ("Ficheiros Excel (antigos)", "*.xls"),
            ("Todos os ficheiros", "*.*"),
        ],
    )
    if not selected:
        return None
    return Path(selected)


def _show_missing_customers_dialog(parent: "tk.Misc", missing_ids: Sequence[str]) -> None:
    import tkinter as tk
    from tkinter import ttk

    window = tk.Toplevel(parent)
    window.title("Clientes em falta no MasterFiles")
    window.transient(parent)
    window.grab_set()
    window.resizable(True, True)

    description = (
        "Foram detectados clientes nas facturas que não existem no MasterFiles. "
        "Indique o ficheiro de clientes que deve ser utilizado para completar os registos."
    )

    label = ttk.Label(window, text=description, wraplength=480, justify="left")
    label.pack(fill="x", padx=16, pady=(16, 8))

    frame = ttk.Frame(window)
    frame.pack(fill="both", expand=True, padx=16, pady=(0, 8))

    columns = max(1, min(4, len(missing_ids)))
    tree_columns = [f"col_{index}" for index in range(columns)]
    tree = ttk.Treeview(
        frame,
        columns=tree_columns,
        show="headings",
        height=min(10, max(1, ceil(len(missing_ids) / columns))),
        selectmode="none",
    )

    for column in tree_columns:
        tree.heading(column, text="Identificador")
        tree.column(column, anchor="center", width=120, stretch=True)

    iterator = iter(missing_ids)
    rows: list[tuple[str, ...]] = []
    while True:
        batch: list[str] = []
        for _ in range(columns):
            try:
                batch.append(next(iterator))
            except StopIteration:
                break
        if not batch:
            break
        while len(batch) < columns:
            batch.append("")
        rows.append(tuple(batch))

    for row in rows:
        tree.insert("", "end", values=row)

    scrollbar = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)

    tree.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")

    button = ttk.Button(window, text="Continuar", command=window.destroy)
    button.pack(pady=(0, 16))

    window.update_idletasks()
    window.minsize(window.winfo_width(), window.winfo_height())
    window.wait_window()


def _prompt_for_customer_file(
    parent: "tk.Misc", *, initialdir: Path | None = None
) -> Path | None:
    from tkinter import filedialog

    selected = filedialog.askopenfilename(
        parent=parent,
        title="Seleccione o ficheiro de clientes",
        initialdir=str(initialdir) if initialdir else None,
        filetypes=[
            ("Ficheiros Excel", "*.xlsx"),
            ("Ficheiros Excel (antigos)", "*.xls"),
            ("Todos os ficheiros", "*.*"),
        ],
    )
    if not selected:
        return None
    return Path(selected)


def _load_records_from_excel(path: Path) -> dict[str, dict[str, str]]:
    from openpyxl import load_workbook

    if not path.exists():
        raise FileNotFoundError(f"Ficheiro Excel não encontrado: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        raise ValueError("O ficheiro Excel não contém dados.")

    header = rows[0]
    column_map = _build_column_map(header)

    records: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        if row is None:
            continue
        customer_id = _normalise_excel_value(_value_at(row, column_map["codigo"]))
        if not customer_id:
            continue
        records[customer_id] = {
            "customer_id": customer_id,
            "company_name": _normalise_excel_value(_value_at(row, column_map["nome"])),
            "tax_id": _normalise_excel_value(
                _value_at(row, column_map["contribuinte"])
            ),
            "city": _normalise_excel_value(_value_at(row, column_map["localidade"])),
            "country": _resolve_country_code(
                _normalise_excel_value(_value_at(row, column_map["pais"]))
            ),
            "telephone": _normalise_excel_value(
                _value_at(row, column_map["telemovel"])
            ),
        }

    if not records:
        raise ValueError(
            "Nenhum registo de cliente válido foi encontrado no ficheiro Excel."
        )
    return records


def _build_column_map(header: tuple[object, ...]) -> dict[str, int]:
    required = {
        "codigo": "Código",
        "nome": "Nome",
        "contribuinte": "Contribuinte",
        "localidade": "Localidade",
        "pais": "País",
        "telemovel": "Telemovel",
    }

    mapping: dict[str, int] = {}
    for index, value in enumerate(header):
        key = _normalise_header(value)
        if key in required and key not in mapping:
            mapping[key] = index

    missing = [orig for key, orig in required.items() if key not in mapping]
    if missing:
        joined = ", ".join(missing)
        raise ValueError(
            "O ficheiro Excel não contém todas as colunas obrigatórias: " f"{joined}."
        )
    return mapping


def _normalise_header(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    normalised = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalised if not unicodedata.combining(ch))


def _normalise_excel_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:g}"
    return str(value).strip()


def _value_at(row: tuple[object, ...], index: int) -> object:
    if index >= len(row):
        return None
    return row[index]


def _ensure_masterfiles_node(root: etree._Element, ns_uri: str) -> etree._Element:
    tag = _ns_tag("MasterFiles", ns_uri)
    masterfiles = root.find(f".//{tag}")
    if masterfiles is not None:
        return masterfiles

    masterfiles = etree.Element(tag)

    source_tag = _ns_tag("SourceDocuments", ns_uri)
    for index, child in enumerate(root):
        if child.tag == source_tag:
            root.insert(index, masterfiles)
            break
    else:
        root.append(masterfiles)
    return masterfiles


def _append_customer(
    masterfiles: etree._Element, ns_uri: str, record: _CustomerRecord
) -> None:
    customer_tag = _ns_tag("Customer", ns_uri)

    insertion_order = {
        _ns_tag("Supplier", ns_uri),
        _ns_tag("Product", ns_uri),
        _ns_tag("TaxTable", ns_uri),
    }

    insert_at = len(masterfiles)
    for index, child in enumerate(masterfiles):
        if child.tag in insertion_order:
            insert_at = index
            break

    nsmap = {
        prefix: uri
        for prefix, uri in (masterfiles.nsmap or {}).items()
        if prefix and uri != ns_uri
    }
    nsmap[None] = ns_uri

    if insert_at == len(masterfiles):
        customer = etree.Element(customer_tag, nsmap=nsmap)
        masterfiles.append(customer)
    else:
        customer = etree.Element(customer_tag, nsmap=nsmap)
        masterfiles.insert(insert_at, customer)

    def add_element(parent: etree._Element, name: str, text: str) -> etree._Element:
        element = etree.SubElement(parent, _ns_tag(name, ns_uri))
        element.text = text
        return element

    add_element(customer, "CustomerID", record.customer_id)
    add_element(customer, "AccountID", record.customer_id)
    add_element(customer, "CustomerTaxID", record.tax_id or "999999990")
    add_element(customer, "CompanyName", record.company_name or record.customer_id)

    billing = etree.SubElement(customer, _ns_tag("BillingAddress", ns_uri))
    add_element(billing, "AddressDetail", record.company_name or "Morada não fornecida")
    add_element(billing, "City", record.city or "Desconhecida")
    if record.country:
        add_element(billing, "Country", record.country)
    else:
        add_element(billing, "Country", "AO")

    if record.telephone:
        add_element(customer, "Telephone", record.telephone)

    add_element(customer, "SelfBillingIndicator", "0")


def _resolve_country_code(raw_value: str) -> str:
    value = (raw_value or "").strip()
    if not value:
        return "AO"

    upper = value.upper()
    if len(upper) == 2 and upper.isalpha():
        return upper

    mapping = _load_country_codes()
    key = _normalise_country_key(value)
    if key in mapping:
        return mapping[key]

    return "AO"


@functools.lru_cache(maxsize=1)
def _load_country_codes() -> dict[str, str]:
    mapping: dict[str, str] = {}
    if not _COUNTRY_CODES_PATH.exists():
        return mapping

    for line in _COUNTRY_CODES_PATH.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped or not stripped.startswith("|"):
            continue
        parts = [part.strip() for part in stripped.strip("|").split("|")]
        if len(parts) < 2:
            continue
        name, code = parts[0], parts[1]
        if not name or not code or name.lower() == "país" or code.lower() == "código":
            continue
        code = code.strip()
        mapping[_normalise_country_key(name)] = code.upper() if len(code) == 2 else code

    return mapping


def _normalise_country_key(value: str) -> str:
    normalised = _normalise_header(value)
    return "".join(ch for ch in normalised if ch.isalnum())


def _ns_tag(name: str, ns_uri: str) -> str:
    return f"{{{ns_uri}}}{name}" if ns_uri else name


def _find_child(
    parent: etree._Element, name: str, ns_uri: str
) -> etree._Element | None:
    return parent.find(_ns_tag(name, ns_uri))


def _find_child_text(parent: etree._Element, name: str, ns_uri: str) -> str:
    child = _find_child(parent, name, ns_uri)
    if child is None or child.text is None:
        return ""
    return child.text.strip()
