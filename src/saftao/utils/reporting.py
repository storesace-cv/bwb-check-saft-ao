"""Helpers to aggregate SAF-T (AO) documents and build Excel reports."""

from __future__ import annotations

import os
from copy import copy
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Iterable

from lxml import etree
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from ..rules import iter_sales_invoices
from ..utils import parse_decimal

REPORT_OUTPUT_ENV = "SAFTAO_REPORT_DIR"
DEFAULT_REPORT_DIR = Path(__file__).resolve().parents[3] / "work" / "destino" / "relatorios"


@dataclass
class Totals:
    """Aggregate of monetary values for a set of documents."""

    net_total: Decimal = field(default_factory=lambda: Decimal("0"))
    tax_total: Decimal = field(default_factory=lambda: Decimal("0"))
    gross_total: Decimal = field(default_factory=lambda: Decimal("0"))
    tax_by_rate: dict[str, Decimal] = field(default_factory=dict)

    def add(
        self,
        net: Decimal,
        tax: Decimal,
        gross: Decimal,
        tax_by_rate: dict[str, Decimal] | None = None,
    ) -> None:
        """Add *net*, *tax* and *gross* values to the totals."""

        self.net_total += net
        self.tax_total += tax
        self.gross_total += gross
        if tax_by_rate:
            for rate, amount in tax_by_rate.items():
                self.tax_by_rate[rate] = self.tax_by_rate.get(rate, Decimal("0")) + amount


@dataclass
class NonAccountingDocument:
    """Representation of a document that does not count towards totals."""

    document_type: str
    document_number: str
    document_date: str
    customer_id: str
    totals: Totals


@dataclass
class ReportData:
    """Container for the aggregated data extracted from a SAF-T file."""

    totals_by_month: dict[str, dict[str, Totals]]
    totals_by_type: dict[str, Totals]
    month_overall_totals: dict[str, Totals]
    overall_totals: Totals
    non_accounting_documents: list[NonAccountingDocument]
    tax_rates: list[str]


@dataclass(frozen=True)
class TemplateStyles:
    month_cell: object
    header_text: object
    header_number: object
    detail_text: object
    detail_number: object
    detail_number_negative: object
    subtotal_text: object
    subtotal_number: object
    row_heights: dict[str, float]
    column_widths: dict[str, float]


def _load_template_styles() -> TemplateStyles | None:
    base_dir = Path(__file__).resolve().parents[3]
    candidates = list((base_dir / "log_files").glob("Template - rel*totais com IVA por tipo.xlsx"))
    if not candidates:
        return None
    template = load_workbook(candidates[0])
    ws = template.active
    return TemplateStyles(
        month_cell=ws["A1"],
        header_text=ws["A2"],
        header_number=ws["B2"],
        detail_text=ws["A3"],
        detail_number=ws["B3"],
        detail_number_negative=ws["B5"],
        subtotal_text=ws["A7"],
        subtotal_number=ws["B7"],
        row_heights={
            "month": ws.row_dimensions[1].height or 24.0,
            "header": ws.row_dimensions[2].height or 21.0,
            "subtotal": ws.row_dimensions[7].height or 21.0,
        },
        column_widths={
            "A": ws.column_dimensions["A"].width or 20.0,
            "B": ws.column_dimensions["B"].width or 18.0,
            "C": ws.column_dimensions["C"].width or 12.0,
            "D": ws.column_dimensions["D"].width or 15.0,
            "E": ws.column_dimensions["E"].width or 18.0,
        },
    )


def resolve_report_directory() -> Path:
    """Return the base directory for generated totals reports."""

    override = os.environ.get(REPORT_OUTPUT_ENV)
    if override:
        return Path(override).expanduser()
    return DEFAULT_REPORT_DIR


def default_report_destination(
    saft_path: Path | None, *, base_dir: Path | None = None
) -> Path:
    """Return the default Excel path for a totals report."""

    directory = base_dir or resolve_report_directory()
    if saft_path is None:
        stem = "relatorio_totais"
    else:
        stem = saft_path.stem or "relatorio_totais"
    destination = directory / f"{stem}_totais.xlsx"
    return destination


def _find_child(element: etree._Element, namespace: str, localname: str) -> etree._Element | None:
    if namespace:
        node = element.find(f"./{{{namespace}}}{localname}")
        if node is not None:
            return node
    for child in element:
        if etree.QName(child).localname == localname:
            return child
    return None


def _find_child_text(element: etree._Element, namespace: str, localname: str) -> str:
    node = _find_child(element, namespace, localname)
    if node is None or node.text is None:
        return ""
    return node.text.strip()


def _extract_document_totals(element: etree._Element, namespace: str) -> Totals:
    totals_node = _find_child(element, namespace, "DocumentTotals")
    if totals_node is None:
        return Totals()

    net = parse_decimal(_find_child_text(totals_node, namespace, "NetTotal"))
    tax = parse_decimal(_find_child_text(totals_node, namespace, "TaxPayable"))
    gross = parse_decimal(_find_child_text(totals_node, namespace, "GrossTotal"))

    totals = Totals()
    totals.add(net, tax, gross)
    return totals


def _format_tax_rate_label(rate: str) -> str:
    if rate == "ND":
        return "IVA-ND"
    return f"IVA-{rate}%"


def _extract_tax_by_rate(element: etree._Element, namespace: str) -> dict[str, Decimal]:
    if namespace:
        tax_nodes = element.xpath(".//n:Tax", namespaces={"n": namespace})
    else:
        tax_nodes = element.findall(".//Tax")
    totals: dict[str, Decimal] = {}
    for tax_node in tax_nodes:
        percentage_text = _find_child_text(tax_node, namespace, "TaxPercentage")
        tax_amount = parse_decimal(_find_child_text(tax_node, namespace, "TaxAmount"))
        if percentage_text:
            rate_key = percentage_text
        else:
            rate_key = "ND" if tax_amount != 0 else "0"
        label = _format_tax_rate_label(rate_key)
        totals[label] = totals.get(label, Decimal("0")) + tax_amount
    return totals


def _resolve_invoice_month(element: etree._Element, namespace: str) -> str:
    date_text = _find_child_text(element, namespace, "InvoiceDate") or ""
    if len(date_text) >= 7 and date_text[4] == "-":
        return date_text[:7]
    return date_text or "Desconhecido"


def _tax_rate_sort_key(label: str) -> tuple[int, Decimal | str]:
    if label == "IVA-ND":
        return (1, "ND")
    value = label.removeprefix("IVA-").removesuffix("%")
    try:
        return (0, Decimal(value))
    except Exception:
        return (1, value)


def _iter_work_documents(root: etree._Element, namespace: str) -> Iterable[etree._Element]:
    if namespace:
        return root.findall(
            ".//n:SourceDocuments/n:WorkingDocuments/n:WorkDocument",
            namespaces={"n": namespace},
        )
    return root.findall(".//SourceDocuments/WorkingDocuments/WorkDocument")


def aggregate_documents(root: etree._Element, namespace: str) -> ReportData:
    """Aggregate accounting and non-accounting documents from *root*."""

    totals_by_month: dict[str, dict[str, Totals]] = {}
    totals_by_type: dict[str, Totals] = {}
    month_overall_totals: dict[str, Totals] = {}
    overall = Totals()
    tax_rates: set[str] = set()

    for invoice in iter_sales_invoices(root, namespace):
        invoice_type = _find_child_text(invoice, namespace, "InvoiceType") or "DESCONHECIDO"
        document_totals = _extract_document_totals(invoice, namespace)
        tax_by_rate = _extract_tax_by_rate(invoice, namespace)
        month = _resolve_invoice_month(invoice, namespace)
        if document_totals.tax_total != 0 and sum(tax_by_rate.values()) == 0 and len(tax_by_rate) == 1:
            rate = next(iter(tax_by_rate))
            tax_by_rate[rate] = document_totals.tax_total
        if invoice_type == "NC":
            document_totals = Totals(
                net_total=-document_totals.net_total,
                tax_total=-document_totals.tax_total,
                gross_total=-document_totals.gross_total,
                tax_by_rate={
                    rate: -amount for rate, amount in tax_by_rate.items()
                },
            )
            tax_by_rate = document_totals.tax_by_rate
        tax_rates.update(tax_by_rate.keys())

        if invoice_type not in totals_by_type:
            totals_by_type[invoice_type] = Totals()
        totals_by_type[invoice_type].add(
            document_totals.net_total,
            document_totals.tax_total,
            document_totals.gross_total,
            tax_by_rate,
        )
        overall.add(
            document_totals.net_total,
            document_totals.tax_total,
            document_totals.gross_total,
            tax_by_rate,
        )

        month_totals = totals_by_month.setdefault(month, {})
        if invoice_type not in month_totals:
            month_totals[invoice_type] = Totals()
        month_totals[invoice_type].add(
            document_totals.net_total,
            document_totals.tax_total,
            document_totals.gross_total,
            tax_by_rate,
        )
        month_overall = month_overall_totals.setdefault(month, Totals())
        month_overall.add(
            document_totals.net_total,
            document_totals.tax_total,
            document_totals.gross_total,
            tax_by_rate,
        )

    non_accounting: list[NonAccountingDocument] = []
    for work_document in _iter_work_documents(root, namespace):
        doc_type = _find_child_text(work_document, namespace, "DocumentType") or "DESCONHECIDO"
        doc_number = _find_child_text(work_document, namespace, "DocumentNumber")
        doc_date = _find_child_text(work_document, namespace, "WorkDate")
        customer = _find_child_text(work_document, namespace, "CustomerID")
        totals = _extract_document_totals(work_document, namespace)
        non_accounting.append(
            NonAccountingDocument(
                document_type=doc_type,
                document_number=doc_number,
                document_date=doc_date,
                customer_id=customer,
                totals=totals,
            )
        )

    return ReportData(
        totals_by_month=totals_by_month,
        totals_by_type=totals_by_type,
        month_overall_totals=month_overall_totals,
        overall_totals=overall,
        non_accounting_documents=non_accounting,
        tax_rates=sorted(tax_rates, key=_tax_rate_sort_key),
    )


def write_excel_report(data: ReportData, destination: Path) -> None:
    """Generate an Excel workbook with accounting totals and other documents."""

    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Resumo"
    tax_columns = data.tax_rates
    header = ["Tipo", "Total sem IVA", *tax_columns, "Total com IVA"]
    template_styles = _load_template_styles()

    def apply_style(cell, template_cell) -> None:
        if template_cell is None:
            return
        cell._style = copy(template_cell._style)
        cell.number_format = template_cell.number_format

    def style_row(row_index: int, row_type: str, *, negative: bool = False) -> None:
        if template_styles is None:
            return
        if row_type == "month":
            template_cell = template_styles.month_cell
            apply_style(summary_ws.cell(row=row_index, column=1), template_cell)
            summary_ws.row_dimensions[row_index].height = template_styles.row_heights["month"]
            return
        if row_type == "header":
            summary_ws.row_dimensions[row_index].height = template_styles.row_heights["header"]
            for col_index in range(1, len(header) + 1):
                template_cell = (
                    template_styles.header_text
                    if col_index == 1
                    else template_styles.header_number
                )
                apply_style(summary_ws.cell(row=row_index, column=col_index), template_cell)
            return
        if row_type == "subtotal":
            summary_ws.row_dimensions[row_index].height = template_styles.row_heights["subtotal"]
            for col_index in range(1, len(header) + 1):
                template_cell = (
                    template_styles.subtotal_text
                    if col_index == 1
                    else template_styles.subtotal_number
                )
                apply_style(summary_ws.cell(row=row_index, column=col_index), template_cell)
            return
        if row_type == "detail":
            for col_index in range(1, len(header) + 1):
                if col_index == 1:
                    template_cell = template_styles.detail_text
                else:
                    template_cell = (
                        template_styles.detail_number_negative
                        if negative
                        else template_styles.detail_number
                    )
                apply_style(summary_ws.cell(row=row_index, column=col_index), template_cell)

    def append_totals_row(label: str, totals: Totals) -> None:
        summary_ws.append(
            [
                label,
                totals.net_total,
                *[totals.tax_by_rate.get(rate, Decimal("0")) for rate in tax_columns],
                totals.gross_total,
            ]
        )
        row_index = summary_ws.max_row
        style_row(row_index, "detail", negative=totals.net_total < 0)

    months = sorted(data.totals_by_month)
    if len(months) > 1:
        for month in months:
            summary_ws.append([f"Mês: {month}"])
            style_row(summary_ws.max_row, "month")
            summary_ws.append(header)
            style_row(summary_ws.max_row, "header")
            for doc_type in sorted(data.totals_by_month[month]):
                append_totals_row(doc_type, data.totals_by_month[month][doc_type])
            append_totals_row(f"Subtotal {month}", data.month_overall_totals[month])
            style_row(summary_ws.max_row, "subtotal")
            summary_ws.append([])
        append_totals_row("Totais Gerais", data.overall_totals)
        style_row(summary_ws.max_row, "subtotal")
    else:
        summary_ws.append(header)
        style_row(summary_ws.max_row, "header")
        for doc_type in sorted(data.totals_by_type):
            append_totals_row(doc_type, data.totals_by_type[doc_type])
        summary_ws.append([])
        append_totals_row("Totais Gerais", data.overall_totals)
        style_row(summary_ws.max_row, "subtotal")

    if template_styles is not None:
        total_columns = len(header)
        for col_index in range(1, total_columns + 1):
            column_letter = get_column_letter(col_index)
            if col_index == 1:
                width = template_styles.column_widths["A"]
            elif col_index == 2:
                width = template_styles.column_widths["B"]
            elif col_index == total_columns:
                width = template_styles.column_widths["E"]
            elif col_index == 3:
                width = template_styles.column_widths["C"]
            else:
                width = template_styles.column_widths["D"]
            summary_ws.column_dimensions[column_letter].width = width

    other_ws = workbook.create_sheet(title="Documentos não contabilísticos")
    other_ws.append(
        [
            "Tipo",
            "Número",
            "Data",
            "Cliente",
            "Total sem IVA",
            "IVA",
            "Total com IVA",
        ]
    )

    for document in data.non_accounting_documents:
        other_ws.append(
            [
                document.document_type,
                document.document_number,
                document.document_date,
                document.customer_id,
                document.totals.net_total,
                document.totals.tax_total,
                document.totals.gross_total,
            ]
        )

    workbook.save(destination)
